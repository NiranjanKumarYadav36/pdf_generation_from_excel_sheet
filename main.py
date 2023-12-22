import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

filepaths = glob.glob("Invoices/*.xlsx")


for filepath in filepaths:

    pdf = FPDF(orientation='p', unit='mm', format='A4')
    pdf.set_auto_page_break(auto='False', margin=0)
    pdf.add_page()

    filename = Path(filepath).stem
    invoice_list = filename.split("-")
    # invoice_nr , date = filename.split("-")

    pdf.set_font(family='Times', style='B', size=15)
    pdf.cell(w=0, h=8, txt=f"Invoice nr. {invoice_list[0]}", ln=1)

    pdf.set_font(family='Times', style='B', size=15)
    pdf.cell(w=0, h=9, txt=f"Order dt. {invoice_list[1]}", ln=1)

    df = pd.read_excel(filepath, sheet_name='Sheet 1')

    columns = list(df.columns)
    columns = [item.replace("_", " ").title() for item in columns]

    pdf.set_font(family='Times', size=10,style='B')
    pdf.set_text_color(0, 0, 0)
    pdf.cell(w=30, h=8, txt=columns[0], border=1)
    pdf.cell(w=70, h=8, txt=columns[1], border=1)
    pdf.cell(w=35, h=8, txt=columns[2], border=1)
    pdf.cell(w=30, h=8, txt=columns[3], border=1)
    pdf.cell(w=30, h=8, txt=columns[4], border=1, ln=1)

    for index, rows in df.iterrows():
        pdf.set_font(family='Times', size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=str(rows['product_id']), border=1)
        pdf.cell(w=70, h=8, txt=str(rows['product_name']), border=1)
        pdf.cell(w=35, h=8, txt=str(rows['amount_purchased']), border=1)
        pdf.cell(w=30, h=8, txt=str(rows['price_per_unit']), border=1)
        pdf.cell(w=30, h=8, txt=str(rows['total_price']), border=1, ln=1)

    total_sum = df["total_price"].sum()

    pdf.set_font(family='Times', size=10)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=70, h=8, txt="", border=1)
    pdf.cell(w=35, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt=str(total_sum), border=1, ln=1)

    pdf.set_font(family='Times', size=10, style='B')
    pdf.cell(w=30, h=8, txt=f"Th total price is {total_sum}", ln=1)

    pdf.set_font(family='Times', size=15, style='B')
    pdf.cell(w=29, h=8, txt=f"PythonHow",)
    pdf.image("pythonhow.png", w=10)


    # dynamic string
    pdf.output(f"PDFs/{filename}.pdf")


